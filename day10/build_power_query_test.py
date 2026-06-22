# Author: bibberzz & Codex, GPT-5.5
# Created: 2026/6/5
# Project: day10
# File: build_power_query_test.py
# Description: 从 users.csv 和 order_deatil.csv 生成适合手动制作透视图和切片器的 power_query_test.xlsx

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent  # 获取当前脚本所在目录
WORKBOOK_PATH = BASE_DIR / "power_query_test.xlsx"  # 最终要生成的 Excel 文件
USERS_CSV = BASE_DIR / "users.csv"  # 用户表 CSV 文件
ORDERS_CSV = BASE_DIR / "order_deatil.csv"  # 订单明细 CSV 文件，文件名按当前本地文件保持不变


def normalize_sex(value):
    """统一性别字段，把 男/男性 统一为 男，把 女/女性 统一为 女。"""
    if pd.isna(value):
        return value

    text = str(value).strip()

    if text in ["男", "男性"]:
        return "男"

    if text in ["女", "女性"]:
        return "女"

    return text


def normalize_addr(value):
    """统一地址字段，把 北京市/广州市 这类末尾的 市 去掉，便于汇总。"""
    if pd.isna(value):
        return value

    text = str(value).strip()

    if text.endswith("市"):
        return text[:-1]

    return text


def load_clean_data():
    """读取 CSV，并完成去重、删除空行、年龄填充、增加星期几、计算订单总金额和用户关联。"""
    users_raw = pd.read_csv(USERS_CSV, encoding="utf-8", na_values=["null", ""])  # 读取 users.csv
    orders_raw = pd.read_csv(ORDERS_CSV, encoding="utf-8", na_values=["null", ""])  # 读取 order_deatil.csv

    users = users_raw.dropna(how="all").drop_duplicates().copy()  # 删除用户表空行和重复行
    orders = orders_raw.dropna(how="all").drop_duplicates().copy()  # 删除订单表空行和重复行

    users["age"] = pd.to_numeric(users["age"], errors="coerce")  # 把 age 转成数值，无法转换的变成 NaN
    age_median = users["age"].median()  # 计算 age 当前列中位数
    users["age"] = users["age"].fillna(age_median).astype(int)  # 用中位数填充 age 空值
    users["sex_clean"] = users["sex"].apply(normalize_sex)  # 增加标准化后的性别列
    users["addr_clean"] = users["addr"].apply(normalize_addr)  # 增加标准化后的地址列

    orders["date"] = pd.to_datetime(orders["date"])  # 把 date 转成日期类型
    orders["id"] = orders["id"].astype(int)  # 把订单 id 转成整数
    orders["user_id"] = orders["user_id"].astype(int)  # 把 user_id 转成整数
    orders["count"] = orders["count"].astype(int)  # 把购买数量转成整数
    orders["order_amount"] = orders["count"] * orders["unit_price"]  # 计算订单总金额

    weekday_map = {
        0: "星期一",
        1: "星期二",
        2: "星期三",
        3: "星期四",
        4: "星期五",
        5: "星期六",
        6: "星期日",
    }

    orders["weekday"] = orders["date"].dt.dayofweek.map(weekday_map)  # 增加星期几字段
    orders["year"] = orders["date"].dt.year  # 增加年份字段，方便做图表

    orders_for_merge = orders.rename(columns={"id": "order_id"})  # 避免订单 id 和用户 id 重名
    users_for_merge = users.rename(columns={"id": "user_id_key"})  # 把用户表 id 临时改名，用于关联

    merged = orders_for_merge.merge(
        users_for_merge,
        left_on="user_id",
        right_on="user_id_key",
        how="left",
    )  # 用 order_deatil.user_id 和 users.id 进行关联
    merged = merged.drop(columns=["user_id_key"])  # 删除临时关联列

    return users_raw, orders_raw, users, orders, merged


def fit_columns(ws):
    """根据内容大致调整列宽。"""
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 32)


def write_dataframe(ws, df, table_name=None):
    """把 DataFrame 写入一个工作表，并可选转换为 Excel 表格。"""
    ws.append(list(df.columns))  # 写入表头

    for _, row in df.iterrows():  # 逐行写入数据
        values = []

        for value in row:
            if pd.isna(value):
                values.append(None)
            else:
                values.append(value)

        ws.append(values)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    fit_columns(ws)

    if table_name and ws.max_row > 1 and ws.max_column > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)


def build_workbook():
    """按照方案 B 生成底表：Python 负责清洗，Excel 负责手动做原生透视图和切片器。"""
    users_raw, orders_raw, users, orders, merged = load_clean_data()

    weekday_order = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    by_weekday = (
        merged.groupby("weekday", as_index=False)["order_amount"]
        .sum()
        .set_index("weekday")
        .reindex(weekday_order)
        .dropna()
        .reset_index()
    )
    by_weekday.columns = ["星期", "订单总金额"]

    by_city = merged.groupby("addr_clean", as_index=False)["order_amount"].sum()
    by_city.columns = ["城市", "订单总金额"]
    by_city = by_city.sort_values("订单总金额", ascending=False)

    by_sex = merged.groupby("sex_clean", as_index=False)["order_amount"].sum()
    by_sex.columns = ["性别", "订单总金额"]
    by_sex = by_sex.sort_values("性别", ascending=False)

    by_name = merged.groupby("name", as_index=False)["order_amount"].sum()
    by_name.columns = ["客户姓名", "订单总金额"]
    by_name = by_name.sort_values("订单总金额", ascending=False)

    wb = Workbook()  # 从零新建一份 Excel
    default_sheet = wb.active
    wb.remove(default_sheet)

    readme_ws = wb.create_sheet("README")  # 创建操作说明工作表
    users_raw_ws = wb.create_sheet("users")  # 创建 users 原始导入表
    orders_raw_ws = wb.create_sheet("order_deatil")  # 创建 order_deatil 原始导入表
    users_ws = wb.create_sheet("users_clean")  # 创建 users 清洗表
    orders_ws = wb.create_sheet("order_detail_clean")  # 创建 order_deatil 清洗表
    merged_ws = wb.create_sheet("order_user_detail")  # 创建订单和用户关联后的明细表
    summary_ws = wb.create_sheet("pivot_reference")  # 创建透视图参考口径工作表

    write_dataframe(users_raw_ws, users_raw, "UsersRaw")
    write_dataframe(orders_raw_ws, orders_raw, "OrderDetailRaw")
    write_dataframe(users_ws, users, "UsersClean")
    write_dataframe(orders_ws, orders, "OrderDetailClean")
    write_dataframe(merged_ws, merged, "OrderUserDetail")

    readme_rows = [
        ["方案 B：Python 生成透视图底表，Excel 手动创建原生透视图和切片器"],
        [""],
        ["已由脚本完成"],
        ["1. 导入 users.csv 和 order_deatil.csv。"],
        ["2. 删除空行和重复行。"],
        ["3. 给 order_deatil 增加 weekday（星期几）和 year。"],
        ["4. 用 order_deatil.user_id 关联 users.id。"],
        ["5. users.age 为空时，用当前 age 列中位数填充。"],
        ["6. 增加 order_amount = count * unit_price，表示订单总金额。"],
        [""],
        ["接下来在 Excel 中手动完成"],
        ["1. 进入 order_user_detail 工作表，确认表格名称是 OrderUserDetail。"],
        ["2. 插入 -> 数据透视表，数据源选择 OrderUserDetail。"],
        ["3. 建 4 个透视表和 4 张透视图："],
        ["   A. 行字段 weekday，值字段 order_amount 求和，标题：探究星期几消费者购买能力最强。"],
        ["   B. 行字段 sex_clean，值字段 order_amount 求和，标题：汇总。"],
        ["   C. 行字段 addr_clean，值字段 order_amount 求和，标题：城市维度销售情况。"],
        ["   D. 行字段 name，值字段 order_amount 求和，按金额降序，标题：客户购买力分析。"],
        ["4. 插入切片器，字段选择 weekday、addr_clean、name、sex_clean。"],
        ["5. 通过切片器的“报表连接/数据透视表连接”，把每个切片器连接到 4 个透视表。"],
        [""],
        ["说明"],
        ["openpyxl 可以生成 xlsx、表格和普通图表，但不能生成 Excel 原生切片器和数据透视表连接。"],
        ["所以这个文件保留最关键的干净底表，原生切片器部分需要在 Excel 里完成；如果 Mac 版 Excel 找不到报表连接，建议用 Windows 版 Excel 完成最后一步。"],
    ]

    for row in readme_rows:
        readme_ws.append(row)

    readme_ws["A1"].font = Font(bold=True, size=16)
    readme_ws["A3"].font = Font(bold=True, size=13)
    readme_ws["A10"].font = Font(bold=True, size=13)
    readme_ws["A21"].font = Font(bold=True, size=13)
    readme_ws.column_dimensions["A"].width = 120

    write_dataframe(summary_ws, by_weekday, "PivotByWeekday")
    start_row = summary_ws.max_row + 3
    for row in [["性别", "订单总金额"], *by_sex.values.tolist()]:
        summary_ws.append(row)
    start_row = summary_ws.max_row + 3
    for row in [["城市", "订单总金额"], *by_city.values.tolist()]:
        summary_ws.append(row)
    start_row = summary_ws.max_row + 3
    for row in [["客户姓名", "订单总金额"], *by_name.values.tolist()]:
        summary_ws.append(row)

    fit_columns(summary_ws)
    wb.save(WORKBOOK_PATH)

    print(f"已生成：{WORKBOOK_PATH}")
    print(f"users 去重清洗后行数：{len(users)}")
    print(f"order_deatil 删除空行后行数：{len(orders)}")
    print(f"关联后明细行数：{len(merged)}")
    print(f"订单总金额合计：{merged['order_amount'].sum():.2f}")


if __name__ == "__main__":
    build_workbook()
